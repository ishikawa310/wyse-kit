"""
統合経営レポート v3
- エグゼクティブサマリ
- 月次PL (会計PL / 経営管理PL 並列表示)
- 現場別収益性ドリルダウン
- 月別売上分解 (B4由来 vs 進行管理由来)
- 月別原価配賦
- 入金予定カレンダー
- 支払予定カレンダー
- 月別売上リスト
- アラート一覧 (重要度別)
"""
import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

DB_PATH = 'wyse.db'
OUTPUT = './経営レポートv3_2025年5月期.xlsx'

FONT = 'Arial'
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
HEADER_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=16, color='1F4E78')
SECTION_FONT = Font(name=FONT, bold=True, size=11, color='1F4E78')
NORMAL = Font(name=FONT, size=10)
BOLD = Font(name=FONT, bold=True, size=10)
MONEY_FMT = '#,##0;(#,##0);-'
PCT_FMT = '0.0%;(0.0%);-'
ALERT_FILL = PatternFill('solid', start_color='F4CCCC')
WARN_FILL = PatternFill('solid', start_color='FFF2CC')
INFO_FILL = PatternFill('solid', start_color='E2EFDA')
OK_FILL = PatternFill('solid', start_color='D9EAD3')
MOTOUKE_FILL = PatternFill('solid', start_color='E2F0D9')
TSUJOU_FILL = PatternFill('solid', start_color='DEEBF7')
TOTAL_FILL = PatternFill('solid', start_color='D9E1F2')
KAIKEI_FILL = PatternFill('solid', start_color='FFF2CC')
KEIEI_FILL = PatternFill('solid', start_color='E2EFDA')

border = Border(left=Side(style='thin', color='BFBFBF'),
                right=Side(style='thin', color='BFBFBF'),
                top=Side(style='thin', color='BFBFBF'),
                bottom=Side(style='thin', color='BFBFBF'))

def setup_headers(ws, headers, row=1, fill=None, font=None):
    fill = fill or HEADER_FILL
    font = font or HEADER_FONT
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    ws.row_dimensions[row].height = 32

def style_money(cell, fmt=MONEY_FMT):
    cell.number_format = fmt
    cell.font = NORMAL
    cell.border = border
    cell.alignment = Alignment(horizontal='right')

def get_data_months(c):
    months = set()
    for r in c.execute("SELECT DISTINCT year, month FROM sales_allocations"):
        months.add((r[0], r[1]))
    for r in c.execute("SELECT DISTINCT year, month FROM cost_allocations"):
        months.add((r[0], r[1]))
    return sorted(months)

def build_summary(wb, c, months):
    ws = wb.active
    ws.title = 'サマリ'
    ws['A1'] = 'ワイズアシスト 経営レポート v3'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f'生成日時: {datetime.now().strftime("%Y-%m-%d %H:%M")} / 会計PL+経営管理PL対応版'
    ws['A2'].font = Font(name=FONT, italic=True, size=10, color='666666')

    # KPI
    n_sites = c.execute('SELECT COUNT(*) FROM sites').fetchone()[0]
    n_motouke = c.execute("SELECT COUNT(*) FROM sites WHERE sales_type='motouke'").fetchone()[0]
    n_alerts_err = c.execute("SELECT COUNT(*) FROM alerts WHERE severity='error'").fetchone()[0]
    n_alerts_warn = c.execute("SELECT COUNT(*) FROM alerts WHERE severity='warning'").fetchone()[0]
    n_alerts_info = c.execute("SELECT COUNT(*) FROM alerts WHERE severity='info'").fetchone()[0]
    total_juchu = c.execute('SELECT SUM(juchu_zeikomi) FROM sites').fetchone()[0] or 0
    total_sales = c.execute('SELECT SUM(sales_amount_zeikomi) FROM sales_allocations').fetchone()[0] or 0
    total_cost_kaikei = c.execute('SELECT SUM(amount_zeibetsu) FROM cost_allocations').fetchone()[0] or 0
    total_cost_keiei = c.execute('SELECT SUM(amount_zeibetsu) FROM cost_allocations_keiei').fetchone()[0] or 0

    ws['A4'] = '【全社KPI】'
    ws['A4'].font = SECTION_FONT
    kpis = [
        ('現場総数', f'{n_sites}件', None),
        ('  うち元請売上', f'{n_motouke}件', None),
        ('  うち通常/解体・造成', f'{n_sites - n_motouke}件', None),
        ('受注金額合計(税込)', f'¥{total_juchu:,.0f}', None),
        ('月別按分合計(税込)', f'¥{total_sales:,.0f}', None),
        ('', '', None),
        ('【会計PL】 (実発生月ベース)', '', None),
        ('原価合計(税別)', f'¥{total_cost_kaikei:,.0f}', None),
        ('粗利合計(税別)', f'¥{total_sales/1.1 - total_cost_kaikei:,.0f}', None),
        ('粗利率', f'{(total_sales/1.1 - total_cost_kaikei)/(total_sales/1.1)*100:.1f}%' if total_sales else 'N/A', None),
        ('', '', None),
        ('【経営管理PL】 (進行率揃え)', '', None),
        ('原価合計(税別)', f'¥{total_cost_keiei:,.0f}', None),
        ('粗利合計(税別)', f'¥{total_sales/1.1 - total_cost_keiei:,.0f}', None),
        ('粗利率', f'{(total_sales/1.1 - total_cost_keiei)/(total_sales/1.1)*100:.1f}%' if total_sales else 'N/A', None),
        ('', '', None),
        ('【データ品質】', '', None),
        ('🔴 エラー', f'{n_alerts_err}件', 'error' if n_alerts_err else None),
        ('🟡 警告', f'{n_alerts_warn}件', 'warning' if n_alerts_warn else None),
        ('ℹ️ 情報', f'{n_alerts_info}件', 'info' if n_alerts_info else None),
    ]
    row = 5
    for k, v, sev in kpis:
        is_header = k.startswith('【') or k.startswith('受注') or k.startswith('現場')
        ws.cell(row=row, column=1, value=k).font = BOLD if is_header else NORMAL
        ws.cell(row=row, column=2, value=v).font = NORMAL
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        if sev == 'error':
            ws.cell(row=row, column=2).fill = ALERT_FILL
        elif sev == 'warning':
            ws.cell(row=row, column=2).fill = WARN_FILL
        elif sev == 'info':
            ws.cell(row=row, column=2).fill = INFO_FILL
        row += 1

    # 直近3ヶ月の月次PL対比
    ws.cell(row=row+1, column=1, value='【直近3ヶ月の月次PL対比】').font = SECTION_FONT
    row += 2
    setup_headers(ws, ['年/月', '売上(税込)',
                       '会計PL原価', '会計粗利率',
                       '経管PL原価', '経管粗利率',
                       '差異(粗利率)'], row=row)
    row += 1
    recent_months = months[-3:] if len(months) >= 3 else months
    for y, m in recent_months:
        s = c.execute("SELECT SUM(sales_amount_zeikomi) FROM sales_allocations WHERE year=? AND month=?", (y, m)).fetchone()[0] or 0
        kc = c.execute("SELECT SUM(amount_zeibetsu) FROM cost_allocations WHERE year=? AND month=?", (y, m)).fetchone()[0] or 0
        kk = c.execute("SELECT SUM(amount_zeibetsu) FROM cost_allocations_keiei WHERE year=? AND month=?", (y, m)).fetchone()[0] or 0
        s_zb = s / 1.1
        kpr = (s_zb - kc) / s_zb if s_zb else 0
        kkpr = (s_zb - kk) / s_zb if s_zb else 0
        ws.cell(row=row, column=1, value=f'{y}/{m:02d}').font = NORMAL
        ws.cell(row=row, column=1).border = border
        for col, val, fmt in [(2, s, MONEY_FMT), (3, kc, MONEY_FMT), (4, kpr, PCT_FMT),
                              (5, kk, MONEY_FMT), (6, kkpr, PCT_FMT),
                              (7, kkpr - kpr, PCT_FMT)]:
            cell = ws.cell(row=row, column=col, value=val)
            style_money(cell, fmt)
        ws.cell(row=row, column=3).fill = KAIKEI_FILL
        ws.cell(row=row, column=4).fill = KAIKEI_FILL
        ws.cell(row=row, column=5).fill = KEIEI_FILL
        ws.cell(row=row, column=6).fill = KEIEI_FILL
        row += 1

    for col, w in [('A', 32), ('B', 18), ('C', 14), ('D', 12), ('E', 14), ('F', 12), ('G', 14)]:
        ws.column_dimensions[col].width = w

def build_pl_dual(wb, c, months):
    ws = wb.create_sheet('月次PL対比')
    ws['A1'] = '月次PL 会計 vs 経営管理 対比'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = '会計PL: 原価を実発生月で計上 / 経営管理PL: 原価を売上の進行率で揃えて計上'
    ws['A2'].font = Font(name=FONT, italic=True, size=9, color='666666')

    # 共通ヘッダ: 売上 / 会計原価 / 会計粗利 / 会計粗利率 / 経管原価 / 経管粗利 / 経管粗利率 / 差異
    setup_headers(ws, ['年/月', '元請売上', '通常解造売上', '売上計(税込)',
                       '会計原価(税別)', '会計粗利', '会計粗利率',
                       '経管原価(税別)', '経管粗利', '経管粗利率',
                       '粗利率差異'], row=4)
    row = 5
    grand = dict(mt=0, ts=0, kc=0, kk=0)
    for y, m in months:
        s_mt = c.execute("""SELECT SUM(a.sales_amount_zeikomi) FROM sales_allocations a JOIN sites s ON a.genba_no=s.genba_no
                            WHERE a.year=? AND a.month=? AND s.sales_type='motouke'""", (y, m)).fetchone()[0] or 0
        s_ts = c.execute("""SELECT SUM(a.sales_amount_zeikomi) FROM sales_allocations a JOIN sites s ON a.genba_no=s.genba_no
                            WHERE a.year=? AND a.month=? AND s.sales_type='tsujou_kaitai'""", (y, m)).fetchone()[0] or 0
        s = s_mt + s_ts
        kc = c.execute("SELECT SUM(amount_zeibetsu) FROM cost_allocations WHERE year=? AND month=?", (y, m)).fetchone()[0] or 0
        kk = c.execute("SELECT SUM(amount_zeibetsu) FROM cost_allocations_keiei WHERE year=? AND month=?", (y, m)).fetchone()[0] or 0
        s_zb = s / 1.1
        gp_k = s_zb - kc
        gp_kk = s_zb - kk
        gpr_k = gp_k / s_zb if s_zb else 0
        gpr_kk = gp_kk / s_zb if s_zb else 0

        ws.cell(row=row, column=1, value=f'{y}/{m:02d}').font = NORMAL
        ws.cell(row=row, column=1).border = border
        cols_data = [(2, s_mt, MONEY_FMT), (3, s_ts, MONEY_FMT), (4, s, MONEY_FMT),
                     (5, kc, MONEY_FMT), (6, gp_k, MONEY_FMT), (7, gpr_k, PCT_FMT),
                     (8, kk, MONEY_FMT), (9, gp_kk, MONEY_FMT), (10, gpr_kk, PCT_FMT),
                     (11, gpr_kk - gpr_k, PCT_FMT)]
        for col, val, fmt in cols_data:
            cell = ws.cell(row=row, column=col, value=val)
            style_money(cell, fmt)
        # 会計/経管の色分け
        for col in (5, 6, 7):
            ws.cell(row=row, column=col).fill = KAIKEI_FILL
        for col in (8, 9, 10):
            ws.cell(row=row, column=col).fill = KEIEI_FILL
        # 差異が大きければハイライト
        if abs(gpr_kk - gpr_k) > 0.10:
            ws.cell(row=row, column=11).fill = WARN_FILL

        grand['mt'] += s_mt; grand['ts'] += s_ts; grand['kc'] += kc; grand['kk'] += kk
        row += 1

    # 合計
    ws.cell(row=row, column=1, value='合計').font = BOLD
    ws.cell(row=row, column=1).fill = TOTAL_FILL
    ws.cell(row=row, column=1).border = border
    g_s = grand['mt'] + grand['ts']
    g_s_zb = g_s / 1.1
    g_gp_k = g_s_zb - grand['kc']
    g_gp_kk = g_s_zb - grand['kk']
    g_gpr_k = g_gp_k / g_s_zb if g_s_zb else 0
    g_gpr_kk = g_gp_kk / g_s_zb if g_s_zb else 0
    cols_data = [(2, grand['mt'], MONEY_FMT), (3, grand['ts'], MONEY_FMT), (4, g_s, MONEY_FMT),
                 (5, grand['kc'], MONEY_FMT), (6, g_gp_k, MONEY_FMT), (7, g_gpr_k, PCT_FMT),
                 (8, grand['kk'], MONEY_FMT), (9, g_gp_kk, MONEY_FMT), (10, g_gpr_kk, PCT_FMT),
                 (11, g_gpr_kk - g_gpr_k, PCT_FMT)]
    for col, val, fmt in cols_data:
        cell = ws.cell(row=row, column=col, value=val)
        style_money(cell, fmt)
        cell.font = BOLD
        cell.fill = TOTAL_FILL

    widths = [12, 14, 16, 14, 14, 14, 10, 14, 14, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'B5'

def build_alerts(wb, c):
    ws = wb.create_sheet('アラート一覧')
    ws['A1'] = 'データ品質アラート一覧'
    ws['A1'].font = TITLE_FONT

    # 重大度別の集計
    ws['A3'] = '【重大度別サマリ】'
    ws['A3'].font = SECTION_FONT
    setup_headers(ws, ['重大度', 'カテゴリ', '件数'], row=4)
    row = 5
    for r in c.execute("""SELECT severity, kind, COUNT(*) FROM alerts
                          GROUP BY severity, kind
                          ORDER BY CASE severity WHEN 'error' THEN 1 WHEN 'warning' THEN 2 ELSE 3 END, kind"""):
        ws.cell(row=row, column=1, value=r[0]).font = BOLD
        ws.cell(row=row, column=1).border = border
        if r[0] == 'error':
            ws.cell(row=row, column=1).fill = ALERT_FILL
        elif r[0] == 'warning':
            ws.cell(row=row, column=1).fill = WARN_FILL
        else:
            ws.cell(row=row, column=1).fill = INFO_FILL
        ws.cell(row=row, column=2, value=r[1]).font = NORMAL
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=3, value=r[2]).font = NORMAL
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
        row += 1

    row += 2
    ws.cell(row=row, column=1, value='【アラート詳細】').font = SECTION_FONT
    row += 1
    setup_headers(ws, ['重大度', 'カテゴリ', '現場No', '現場名', 'メッセージ', '詳細'], row=row)
    row += 1
    for r in c.execute("""SELECT a.severity, a.kind, a.genba_no, s.genba_name, a.message, a.detail
                          FROM alerts a LEFT JOIN sites s ON a.genba_no=s.genba_no
                          ORDER BY CASE a.severity WHEN 'error' THEN 1 WHEN 'warning' THEN 2 ELSE 3 END,
                          a.kind, a.genba_no"""):
        for j, v in enumerate(r):
            cell = ws.cell(row=row, column=j+1, value=v)
            cell.font = NORMAL
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        if r[0] == 'error':
            ws.cell(row=row, column=1).fill = ALERT_FILL
        elif r[0] == 'warning':
            ws.cell(row=row, column=1).fill = WARN_FILL
        else:
            ws.cell(row=row, column=1).fill = INFO_FILL
        row += 1

    widths = [10, 22, 10, 25, 38, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A5'

def build_site_profitability(wb, c):
    ws = wb.create_sheet('現場別収益性')
    ws['A1'] = '現場別収益性ドリルダウン'
    ws['A1'].font = TITLE_FONT
    setup_headers(ws, ['現場No', '現場名', '発注者', '売上種別', '工事区分', '完了',
                       '請負金額(税込)', '採用原価(税別)', '粗利', '粗利率',
                       'B4状態', '計上売上累計', '入金予定日'], row=3)
    row = 4
    sales_type_map = {'motouke': '元請', 'tsujou_kaitai': '通常/解体造成'}
    b4_label = {'ok': 'B4採用', 'empty': 'B4空欄', 'unparseable': 'B4解析失敗',
                'no_sheet': '現場シートなし'}
    for r in c.execute("""
        SELECT s.genba_no, s.genba_name, s.hatchusha, s.sales_type, s.koji_kubun, s.is_completed,
               s.juchu_zeikomi, s.juchu_zeinuki,
               s.shitauke_zeibetsu, s.genka_yotei, s.genka_jissai, s.hendouhi_shinkou,
               s.b4_status, s.nyukin_yotei_date,
               (SELECT SUM(sales_amount_zeikomi) FROM sales_allocations WHERE genba_no=s.genba_no) as sales_alloc
        FROM sites s ORDER BY s.genba_no
    """):
        (no, nm, h, st, kk, cmpl, juchu_zk, juchu_zb, shitauke, gy, gj, hendou,
         b4st, nyukin, sales_alloc) = r
        if st == 'motouke':
            adopted = shitauke or 0
        else:
            adopted = (gj if cmpl else None) or hendou or 0
        juchu_zb = juchu_zb or (juchu_zk / 1.1 if juchu_zk else 0)
        gp = juchu_zb - adopted
        gpr = gp / juchu_zb if juchu_zb else 0

        ws.cell(row=row, column=1, value=no).font = NORMAL
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2, value=nm).font = NORMAL
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=3, value=h).font = NORMAL
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=4, value=sales_type_map.get(st, st)).font = NORMAL
        ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=4).fill = MOTOUKE_FILL if st == 'motouke' else TSUJOU_FILL
        ws.cell(row=row, column=5, value=kk).font = NORMAL
        ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=6, value='完了' if cmpl else '未完了').font = NORMAL
        ws.cell(row=row, column=6).border = border
        for col, val, fmt in [(7, juchu_zk, MONEY_FMT), (8, adopted, MONEY_FMT),
                              (9, gp, MONEY_FMT), (10, gpr, PCT_FMT)]:
            cell = ws.cell(row=row, column=col, value=val)
            style_money(cell, fmt)
        if gp < 0:
            ws.cell(row=row, column=9).fill = ALERT_FILL
            ws.cell(row=row, column=10).fill = ALERT_FILL
        ws.cell(row=row, column=11, value=b4_label.get(b4st, b4st)).font = NORMAL
        ws.cell(row=row, column=11).border = border
        if b4st == 'unparseable':
            ws.cell(row=row, column=11).fill = WARN_FILL
        elif b4st in ('empty', 'no_sheet'):
            ws.cell(row=row, column=11).fill = INFO_FILL
        style_money(ws.cell(row=row, column=12, value=sales_alloc))
        ws.cell(row=row, column=13, value=nyukin).font = NORMAL
        ws.cell(row=row, column=13).border = border
        row += 1

    widths = [10, 25, 22, 14, 14, 8, 16, 16, 14, 10, 14, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'C4'

def build_nyukin_calendar(wb, c):
    ws = wb.create_sheet('入金予定')
    ws['A1'] = '入金予定カレンダー'
    ws['A1'].font = TITLE_FONT
    setup_headers(ws, ['現場No', '現場名', '売上種別', '請負金額(税込)', '入金予定日', '計上累計'], row=3)
    row = 4
    sales_type_map = {'motouke': '元請', 'tsujou_kaitai': '通常/解体造成'}
    for r in c.execute("""
        SELECT s.genba_no, s.genba_name, s.sales_type, s.juchu_zeikomi, s.nyukin_yotei_date,
               (SELECT SUM(sales_amount_zeikomi) FROM sales_allocations WHERE genba_no=s.genba_no) as sa
        FROM sites s WHERE s.nyukin_yotei_date != '' AND s.juchu_zeikomi > 0
        ORDER BY s.nyukin_yotei_date
    """):
        ws.cell(row=row, column=1, value=r[0]).font = NORMAL
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2, value=r[1]).font = NORMAL
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=3, value=sales_type_map.get(r[2], r[2])).font = NORMAL
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=3).fill = MOTOUKE_FILL if r[2] == 'motouke' else TSUJOU_FILL
        style_money(ws.cell(row=row, column=4, value=r[3]))
        ws.cell(row=row, column=5, value=r[4]).font = BOLD
        ws.cell(row=row, column=5).border = border
        style_money(ws.cell(row=row, column=6, value=r[5]))
        row += 1
    # 未入力数を表示
    no_nyukin = c.execute("SELECT COUNT(*) FROM sites WHERE nyukin_yotei_date='' AND juchu_zeikomi > 0").fetchone()[0]
    ws.cell(row=row+1, column=1, value=f'※ 入金予定日未入力: {no_nyukin}件 (アラート一覧シート参照)').font = Font(name=FONT, italic=True, size=9, color='C00000')

    for col, w in [('A',10),('B',25),('C',14),('D',16),('E',14),('F',16)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A4'

def build_shiharai_calendar(wb, c):
    ws = wb.create_sheet('支払予定')
    ws['A1'] = '支払予定カレンダー(原価明細)'
    ws['A1'].font = TITLE_FONT
    setup_headers(ws, ['支払日', '現場No', '現場名', '費目', '取引先', '金額(税別)', '金額(税込)'], row=3)
    row = 4
    for r in c.execute("""
        SELECT co.shiharai_date, co.genba_no, s.genba_name, co.hi_moku, co.torihiki_saki,
               co.kingaku_zeibetsu, co.kingaku_zeikomi
        FROM costs co JOIN sites s ON co.genba_no=s.genba_no
        WHERE co.shiharai_date != '' ORDER BY co.shiharai_date
    """):
        for col, v in enumerate(r[:5], 1):
            ws.cell(row=row, column=col, value=v).font = NORMAL
            ws.cell(row=row, column=col).border = border
        style_money(ws.cell(row=row, column=6, value=r[5]))
        style_money(ws.cell(row=row, column=7, value=r[6]))
        row += 1

    for col, w in [('A',12),('B',10),('C',25),('D',12),('E',22),('F',14),('G',14)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A4'

def build_sales_list(wb, c):
    ws = wb.create_sheet('月別売上リスト')
    setup_headers(ws, ['現場No', '現場名', '請負金額(税込)', '下請金額(税抜)', '売上種別',
                       '年', '月', '割合(割)', '売上金額(税込)', '下請金額按分(税込)', '計上ソース'], row=1)
    row = 2
    sales_type_map = {'motouke': '元請売上', 'tsujou_kaitai': '通常/解体造成'}
    for r in c.execute("""
        SELECT a.genba_no, s.genba_name, s.juchu_zeikomi, s.shitauke_zeibetsu, s.sales_type,
               a.year, a.month, a.ratio_wari, a.sales_amount_zeikomi, a.shitauke_amount_zeikomi, a.source
        FROM sales_allocations a JOIN sites s ON a.genba_no=s.genba_no
        ORDER BY a.year, a.month, s.sales_type, a.genba_no
    """):
        vals = list(r)
        vals[4] = sales_type_map.get(vals[4], vals[4])
        for j, v in enumerate(vals):
            cell = ws.cell(row=row, column=j+1, value=v)
            cell.font = NORMAL
            cell.border = border
            if j in (2, 3, 8, 9):
                cell.number_format = MONEY_FMT
            elif j == 7:
                cell.number_format = '0.00'
        fill = MOTOUKE_FILL if vals[4] == '元請売上' else TSUJOU_FILL
        for jj in range(1, 12):
            ws.cell(row=row, column=jj).fill = fill
        row += 1
    widths = [10, 25, 16, 14, 14, 8, 6, 10, 16, 18, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

def main():
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    months = get_data_months(c)

    wb = Workbook()
    build_summary(wb, c, months)
    build_pl_dual(wb, c, months)
    build_alerts(wb, c)
    build_site_profitability(wb, c)
    build_nyukin_calendar(wb, c)
    build_shiharai_calendar(wb, c)
    build_sales_list(wb, c)

    wb.save(OUTPUT)
    print(f'保存: {OUTPUT}')
    conn.close()

if __name__ == '__main__':
    main()
