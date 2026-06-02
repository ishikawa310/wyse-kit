"""
ワイズアシスト現場管理表 抽出パイプライン v2
進行管理シート + 全現場シートから5テーブル(sites, sales_allocations, costs, cost_allocations, alerts)に正規化
v2: 現場シートB4「売上計上月」を売上計上の一次源とし、進行管理との差分をアラートとして検出
"""
import re
import sqlite3
import unicodedata
import calendar as _cal
from pathlib import Path
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from parse_b4 import parse_uriage_tsuki
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

_EXCEL_EPOCH = date(1899, 12, 30)  # Excelシリアル日付の起点

def _norm(s):
    """全角→半角・前後空白除去（シート名比較用）"""
    return unicodedata.normalize('NFKC', str(s)).strip()

# ===== 設定 =====
BLUE_COLORS = {'FF4A86E8', 'FF6D9EEB', 'FF6FA8DC'}
RED_COLORS = {'FFEA9999'}
PINK_EXCLUDE = {'FFF4CCCC'}
BILLING_COLORS = BLUE_COLORS | RED_COLORS

EXCEL_PATH = 'genba.xlsx'
DB_PATH = 'wyse.db'

# ===== ユーティリティ =====
def get_color(cell):
    """セルの塗りつぶし色をARGBで取得"""
    fill = cell.fill
    if not fill or not fill.fgColor:
        return None
    rgb = fill.fgColor.rgb
    if rgb and rgb != '00000000':
        return rgb
    return None

def extract_number(val):
    """テキスト混じり数値から数値部分を抽出(例: '（仮）20,000,000' -> 20000000)"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(',', '')
    m = re.search(r'\d+(?:\.\d+)?', s)
    return float(m.group()) if m else None

def parse_date(val):
    """セル値を日付に変換（datetime/date/文字列/Excelシリアル値に対応）"""
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # 文字列: "2026/09/30", "2026-09-30", "2026/9/30" など
    if isinstance(val, str):
        for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%Y/%m', '%m/%d/%Y'):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    # 数値: openpyxlがExcelシリアル日付をfloat/intで返す場合
    if isinstance(val, (int, float)):
        try:
            return _EXCEL_EPOCH + timedelta(days=int(val))
        except (ValueError, OverflowError):
            pass
    return None

# ===== 進行管理シート抽出 =====
def extract_shinkou_kanri(wb_color, wb_data):
    """進行管理シートから月別売上按分を生成。
    返却: list of dict
    """
    ws_c = wb_color['進行管理']  # 色判定用
    ws_d = wb_data['進行管理']    # 値用

    YEAR_ROW = 1
    MONTH_ROW = 2
    DATA_START_ROW = 4
    MONTH_COL_START = 17  # Q列
    MONTH_COL_END = 52    # 2026年12月まで

    # 月カラムのマップ作成(年は飛び飛びなので前方補完、文字列含む年表記にも対応)
    import re as _re
    month_cols = []
    current_year = None
    for c in range(MONTH_COL_START, MONTH_COL_END + 1):
        year_v = ws_d.cell(row=YEAR_ROW, column=c).value
        month_v = ws_d.cell(row=MONTH_ROW, column=c).value
        if year_v is not None:
            # 数値直接 or "2026~" のような文字列から年を抽出
            if isinstance(year_v, (int, float)):
                current_year = int(year_v)
            else:
                m = _re.search(r'(20\d{2})', str(year_v))
                if m:
                    current_year = int(m.group(1))
        if month_v and current_year:
            try:
                current_year_tmp = current_year
                month_int = int(month_v)
                # 年跨ぎ検出: 月が1〜3で前の月が10〜12なら年++(月ヘッダ自体が連続している前提が崩れた時の保険)
                if month_cols and month_cols[-1][2] >= 10 and month_int <= 3 and month_cols[-1][1] == current_year_tmp:
                    # ヘッダで年が更新されていない場合のみ補正
                    yv_next = ws_d.cell(row=YEAR_ROW, column=c).value
                    if yv_next is None:
                        current_year_tmp = current_year + 1
                        current_year = current_year_tmp
                month_cols.append((c, current_year_tmp, month_int))
            except (ValueError, TypeError):
                pass

    sites = []
    excluded = []
    for r in range(DATA_START_ROW, ws_d.max_row + 1):
        genba_no = ws_d.cell(row=r, column=1).value
        if not genba_no:
            continue
        genba_no = str(genba_no).strip()
        # 行内に '全体' のような集計行をスキップ
        if genba_no.lower() in ('000', 'nan'):
            continue
        genba_name = ws_d.cell(row=r, column=2).value
        juchu_zeinuki = ws_d.cell(row=r, column=3).value     # C列 請負金額(税抜)
        juchu_zeikomi = ws_d.cell(row=r, column=4).value     # D列 請負金額(税込)
        genka_mitsumori = ws_d.cell(row=r, column=5).value   # E列 原価見積額
        shitauke_raw = ws_d.cell(row=r, column=6).value      # F列 下請金額
        chakkin_date_raw = ws_d.cell(row=r, column=7).value  # G列 着金見込み日(テキスト混じり)
        hendouhi = ws_d.cell(row=r, column=8).value          # H列 変動費(原価)
        shiharai_date_raw = ws_d.cell(row=r, column=9).value # I列 支払日(テキスト混じり)
        sesshu_name = ws_d.cell(row=r, column=10).value      # J列 施主名(元請)
        shokaisha = ws_d.cell(row=r, column=11).value        # K列 紹介者
        is_completed_flag = ws_d.cell(row=r, column=12).value # L列 現場終了

        # 除外: 請負金額未入力or0
        juchu = extract_number(juchu_zeikomi)
        if not juchu or juchu == 0:
            excluded.append((genba_no, genba_name, '請負金額未入力/0'))
            continue

        # 計上月を特定
        billing_months = []
        for col, year, month in month_cols:
            cell_c = ws_c.cell(row=r, column=col)
            color = get_color(cell_c)
            if color in BILLING_COLORS:
                ratio_val = ws_d.cell(row=r, column=col).value
                billing_months.append((year, month, ratio_val))

        if not billing_months:
            excluded.append((genba_no, genba_name, '青/赤セルなし'))
            continue

        # 割合決定
        with_values = [(y, m, extract_number(rv)) for y, m, rv in billing_months if extract_number(rv)]
        if with_values and len(with_values) == len(billing_months):
            # すべて数値あり
            allocations_ratio = [(y, m, v) for y, m, v in with_values]
        elif not with_values:
            # すべて空白 -> 均等分割
            per = 10.0 / len(billing_months)
            allocations_ratio = [(y, m, per) for y, m, _ in billing_months]
        else:
            # 混在 -> 数値ありはそのまま、残りは(10 - 数値合計)/空白数 で按分
            sum_val = sum(v for _, _, v in with_values)
            remain = 10.0 - sum_val
            blank_count = len(billing_months) - len(with_values)
            per = remain / blank_count if blank_count > 0 else 0
            allocations_ratio = []
            for y, m, rv in billing_months:
                v = extract_number(rv)
                if v is not None:
                    allocations_ratio.append((y, m, v))
                else:
                    allocations_ratio.append((y, m, per))

        # 売上種別
        shitauke_zeibetsu = extract_number(shitauke_raw)
        sales_type = 'motouke' if shitauke_zeibetsu else 'tsujou_kaitai'

        # 売上金額計算
        allocations = []
        for y, m, ratio in allocations_ratio:
            sales_amt = juchu * ratio / 10
            shitauke_amt = None
            if sales_type == 'motouke' and shitauke_zeibetsu:
                # 下請按分(税込) = 税抜 × 1.1 × 割合/10
                shitauke_amt = shitauke_zeibetsu * 1.1 * ratio / 10
            allocations.append({
                'year': y, 'month': m, 'ratio': ratio,
                'sales_amount_zeikomi': sales_amt,
                'shitauke_amount_zeikomi': shitauke_amt
            })

        sites.append({
            'genba_no': genba_no,
            'genba_name': str(genba_name) if genba_name else '',
            'juchu_zeikomi': juchu,
            'juchu_zeinuki': extract_number(juchu_zeinuki),
            'genka_mitsumori_shinkou': extract_number(genka_mitsumori),
            'hendouhi_shinkou': extract_number(hendouhi),
            'shitauke_zeibetsu': shitauke_zeibetsu,
            'sales_type': sales_type,
            'is_completed_shinkou': bool(is_completed_flag) if is_completed_flag is not None else None,
            'chakkin_mikomi_text': str(chakkin_date_raw) if chakkin_date_raw else '',
            'shiharai_text_shinkou': str(shiharai_date_raw) if shiharai_date_raw else '',
            'sesshu_name': str(sesshu_name) if sesshu_name else '',
            'shokaisha': str(shokaisha) if shokaisha else '',
            'allocations_shinkou': allocations,  # 進行管理由来(参考保持)
            'allocations': allocations,  # 一次的にはこれ。後段でB4から上書き可能
        })

    return sites, excluded

# ===== 現場シート抽出 =====
def identify_template(ws):
    """テンプレート種別を判定"""
    e7 = str(ws.cell(row=7, column=5).value or '')
    a14 = str(ws.cell(row=14, column=1).value or '')
    if '予定原価' in e7:
        return 'shin_genka'  # 新原価管理(通常売上)
    if 'アスベスト' in str(ws.cell(row=14, column=2).value or ''):
        return 'motouke'     # 元請売上テンプレート
    if '労務費' in a14:
        return 'kyu_genka'   # 旧原価管理(通常売上)
    return 'unknown'

def find_sections(ws):
    """セクションラベルの位置を動的に検出"""
    sections = {}
    section_labels = {
        '労務費': 'roumu',
        '機材レンタル': 'kizai',
        '処分費': 'shobun',
        '外注費': 'gaicyuhi',
        '下請外注費': 'shitauke',
        '下請外注': 'shitauke',
        '仕入': 'shiire',
        '燃料': 'nenryo',
        'アスベスト事前調査費': 'asbestos',
        'CF管理': 'cf_kanri',
        '【CF管理項目】': 'cf_kanri',
    }
    for r in range(20, min(ws.max_row + 1, 200)):
        label = ws.cell(row=r, column=1).value
        if not label:
            continue
        label_str = str(label).strip()
        for key, name in section_labels.items():
            if label_str.startswith(key) and name not in sections:
                sections[name] = r
                break
    return sections

def extract_site_sheet(ws, template):
    """個別現場シートから財務サマリと原価明細を抽出"""
    info = {
        'chakkou_date': parse_date(ws.cell(row=2, column=1).value),
        'genba_name_sheet': ws.cell(row=2, column=3).value,  # 現場シート短縮名(進行管理が一次源)
        'hatchusha': ws.cell(row=2, column=4).value,
        'ichijuke': ws.cell(row=2, column=6).value,
        'uriage_tsuki_text': ws.cell(row=5, column=1).value,
        'nyukin_yotei_date': parse_date(ws.cell(row=5, column=2).value),
        'koji_kubun': ws.cell(row=5, column=8).value,
        'jucyu_jokyou': ws.cell(row=11, column=8).value,
        'sekoujoukyou': ws.cell(row=11, column=9).value,
    }

    info['juchu_zeibetsu'] = extract_number(ws.cell(row=8, column=1).value)
    info['shohizei'] = extract_number(ws.cell(row=8, column=2).value)
    info['juchu_plus_tsuika'] = extract_number(ws.cell(row=8, column=4).value)
    info['tsuika_koji_zeibetsu'] = extract_number(ws.cell(row=11, column=1).value)

    if template == 'shin_genka':
        info['genka_yotei'] = extract_number(ws.cell(row=8, column=5).value)
        info['rieki_yotei'] = extract_number(ws.cell(row=8, column=6).value)
        info['rieki_ritsu_yotei'] = extract_number(ws.cell(row=8, column=7).value)
        info['genka_jissai'] = extract_number(ws.cell(row=8, column=8).value)
        info['rieki_jissai'] = extract_number(ws.cell(row=8, column=9).value)
        info['rieki_ritsu_jissai'] = extract_number(ws.cell(row=8, column=10).value)
    else:
        info['genka_yotei'] = None
        info['genka_jissai'] = extract_number(ws.cell(row=8, column=5).value)
        info['rieki_jissai'] = extract_number(ws.cell(row=8, column=6).value)
        info['rieki_ritsu_jissai'] = extract_number(ws.cell(row=8, column=7).value)

    if template == 'motouke':
        info['cost_breakdown'] = {
            'gaicyuhi': extract_number(ws.cell(row=15, column=1).value),
            'asbestos': extract_number(ws.cell(row=15, column=2).value),
        }
    else:
        info['cost_breakdown'] = {
            'roumu': extract_number(ws.cell(row=15, column=1).value),
            'kizai_rental': extract_number(ws.cell(row=15, column=2).value),
            'shobun': extract_number(ws.cell(row=15, column=3).value),
            'gaicyuhi': extract_number(ws.cell(row=15, column=4).value),
            'shiire': extract_number(ws.cell(row=15, column=5).value),
            'nenryo': extract_number(ws.cell(row=15, column=6).value),
        }

    # セクション位置を動的検出
    sections = find_sections(ws)

    # セクションの順序(終了位置を決めるため)
    section_order = ['roumu', 'kizai', 'shobun', 'shitauke', 'gaicyuhi', 'shiire', 'nenryo', 'asbestos', 'cf_kanri']
    section_rows = sorted([(name, r) for name, r in sections.items() if name in section_order], key=lambda x: x[1])

    def get_section_range(name):
        """セクションのデータ行範囲を取得"""
        if name not in sections:
            return None, None
        start = sections[name] + 2  # ヘッダ行の次から
        end = None
        for n, r in section_rows:
            if r > sections[name]:
                end = r - 1
                break
        if end is None:
            end = min(sections[name] + 30, ws.max_row)
        return start, end

    costs = []

    # 労務費
    s, e = get_section_range('roumu')
    if s:
        for r in range(s, e + 1):
            d = parse_date(ws.cell(row=r, column=1).value)
            kn = extract_number(ws.cell(row=r, column=2).value)
            ki = extract_number(ws.cell(row=r, column=3).value)
            keisan = extract_number(ws.cell(row=r, column=4).value)
            if d and keisan and keisan > 0:
                costs.append({
                    'hi_moku': '労務費', 'hassei_date': d,
                    'kingaku_zeibetsu': keisan, 'kingaku_zeikomi': keisan,
                    'memo': f'人数{kn} 日当{ki}' if kn and ki else ''
                })

    # 機材レンタル
    s, e = get_section_range('kizai')
    if s:
        for r in range(s, e + 1):
            seikyu_d = parse_date(ws.cell(row=r, column=1).value)
            torihiki = ws.cell(row=r, column=2).value
            shiyo_tsuki = ws.cell(row=r, column=3).value
            tsuki_gokei = extract_number(ws.cell(row=r, column=4).value)
            shiharai_d = parse_date(ws.cell(row=r, column=5).value)
            # 取引先が空 かつ 金額もない行はスキップ
            if not (torihiki or tsuki_gokei):
                continue
            if tsuki_gokei and tsuki_gokei > 0:
                costs.append({
                    'hi_moku': '機材レンタル',
                    'torihiki_saki': str(torihiki) if torihiki else '',
                    'hassei_date': seikyu_d,
                    'shiharai_date': shiharai_d,
                    'hassei_tsuki': str(shiyo_tsuki) if shiyo_tsuki else '',
                    'kingaku_zeibetsu': tsuki_gokei,
                    'kingaku_zeikomi': tsuki_gokei * 1.1,
                })

    # 処分費
    s, e = get_section_range('shobun')
    if s:
        for r in range(s, e + 1):
            d = parse_date(ws.cell(row=r, column=1).value)
            torihiki = ws.cell(row=r, column=2).value
            keisan = extract_number(ws.cell(row=r, column=6).value)
            if d and keisan and keisan > 0:
                costs.append({
                    'hi_moku': '処分費',
                    'torihiki_saki': str(torihiki) if torihiki else '',
                    'hassei_date': d,
                    'kingaku_zeibetsu': keisan,
                    'kingaku_zeikomi': keisan * 1.1,
                })

    # 外注費(通常売上で出てくる)
    s, e = get_section_range('gaicyuhi')
    if s:
        for r in range(s, e + 1):
            d = parse_date(ws.cell(row=r, column=1).value)
            torihiki = ws.cell(row=r, column=2).value
            keisan = extract_number(ws.cell(row=r, column=6).value)
            if d and keisan and keisan > 0:
                costs.append({
                    'hi_moku': '外注費',
                    'torihiki_saki': str(torihiki) if torihiki else '',
                    'hassei_date': d,
                    'kingaku_zeibetsu': keisan,
                    'kingaku_zeikomi': keisan * 1.1,
                })

    # 下請外注(元請売上)
    s, e = get_section_range('shitauke')
    if s:
        for r in range(s, e + 1):
            d = parse_date(ws.cell(row=r, column=1).value)
            torihiki = ws.cell(row=r, column=2).value
            keisan = extract_number(ws.cell(row=r, column=6).value)
            shiharai = parse_date(ws.cell(row=r, column=8).value)
            if (torihiki or (keisan and keisan > 0)) and keisan and keisan > 0:
                costs.append({
                    'hi_moku': '下請外注',
                    'torihiki_saki': str(torihiki) if torihiki else '',
                    'hassei_date': d,
                    'shiharai_date': shiharai,
                    'kingaku_zeibetsu': keisan,
                    'kingaku_zeikomi': keisan * 1.1,
                })

    # 仕入
    s, e = get_section_range('shiire')
    if s:
        for r in range(s, e + 1):
            d = parse_date(ws.cell(row=r, column=1).value)
            torihiki = ws.cell(row=r, column=2).value
            keisan = extract_number(ws.cell(row=r, column=6).value)
            if d and keisan and keisan > 0:
                costs.append({
                    'hi_moku': '仕入',
                    'torihiki_saki': str(torihiki) if torihiki else '',
                    'hassei_date': d,
                    'kingaku_zeibetsu': keisan,
                    'kingaku_zeikomi': keisan * 1.1,
                })

    # 燃料/高速代
    s, e = get_section_range('nenryo')
    if s:
        for r in range(s, e + 1):
            d = parse_date(ws.cell(row=r, column=1).value)
            keisan = extract_number(ws.cell(row=r, column=6).value)
            if d and keisan and keisan > 0:
                costs.append({
                    'hi_moku': '燃料高速',
                    'hassei_date': d,
                    'kingaku_zeibetsu': keisan,
                    'kingaku_zeikomi': keisan * 1.1,
                })

    # アスベスト
    s, e = get_section_range('asbestos')
    if s:
        for r in range(s, e + 1):
            gyosha = ws.cell(row=r, column=2).value
            kingaku = extract_number(ws.cell(row=r, column=3).value)
            hattyu = parse_date(ws.cell(row=r, column=6).value)
            shiharai = parse_date(ws.cell(row=r, column=7).value)
            if kingaku and kingaku > 0:
                costs.append({
                    'hi_moku': 'アスベスト',
                    'torihiki_saki': str(gyosha) if gyosha else '',
                    'hassei_date': hattyu,
                    'shiharai_date': shiharai,
                    'kingaku_zeibetsu': kingaku,
                    'kingaku_zeikomi': kingaku * 1.1,
                })

    info['costs'] = costs
    info['_sections'] = sections
    return info

# ===== 入金予定シート抽出 =====
def extract_nyukin_yotei(wb_data):
    """「入金予定」シートから入金スケジュールを抽出。
    列構成: A=現場No, B=現場名, C=請求種別, D=入金金額(税込), E=入金予定日
    """
    sheet_norm_map = {_norm(n): n for n in wb_data.sheetnames}
    actual = sheet_norm_map.get(_norm('入金予定'))
    if not actual:
        print("  [入金予定シートなし] CFカレンダーは sites.nyukin_yotei_date を使用")
        return []

    ws = wb_data[actual]
    _mon_abbr = {v.lower(): k for k, v in enumerate(_cal.month_abbr) if v}
    payments = []

    for r in range(2, ws.max_row + 1):
        genba_no_val = ws.cell(row=r, column=1).value
        if not genba_no_val:
            continue
        genba_no = _norm(str(genba_no_val))
        if not genba_no or genba_no.lower() == 'none':
            continue

        genba_name   = ws.cell(row=r, column=2).value
        seikyu_kubun = ws.cell(row=r, column=3).value
        kingaku_val  = ws.cell(row=r, column=4).value
        nyukin_val   = ws.cell(row=r, column=5).value

        kingaku = extract_number(kingaku_val)
        if not kingaku or kingaku <= 0:
            continue

        # 入金予定日の解析
        nyukin_date = parse_date(nyukin_val)
        nyukin_text = ''
        is_collected = 0

        if nyukin_date is None and nyukin_val is not None:
            raw = str(nyukin_val).strip()
            nyukin_text = raw
            # 受領済・相殺フラグ
            if any(kw in raw for kw in ('受領済', '入金済', '相殺', '済')):
                is_collected = 1
            # YYYY/MM/DD または YYYY-MM-DD を文字列中から抽出
            m = re.search(r'(\d{4})[/\-年](\d{1,2})[/\-月](\d{1,2})', raw)
            if m:
                try:
                    nyukin_date = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                except ValueError:
                    pass
            # "Jan-26", "Mar-26" 形式 (MMM-YY) → その月末
            if nyukin_date is None:
                mm = re.match(r'([A-Za-z]{3})[- ](\d{2})$', raw)
                if mm:
                    mon = _mon_abbr.get(mm.group(1).lower())
                    if mon:
                        yr  = 2000 + int(mm.group(2))
                        nyukin_date = date(yr, mon, _cal.monthrange(yr, mon)[1])
            # "3月末", "4月末" 形式 → 当年のその月末
            if nyukin_date is None:
                mm = re.match(r'(\d{1,2})月末', raw)
                if mm:
                    mon = int(mm.group(1))
                    yr  = datetime.now().year
                    nyukin_date = date(yr, mon, _cal.monthrange(yr, mon)[1])

        payments.append({
            'genba_no':          genba_no,
            'genba_name':        str(genba_name).strip() if genba_name else '',
            'seikyu_kubun':      str(seikyu_kubun).strip() if seikyu_kubun else '',
            'kingaku_zeikomi':   kingaku,
            'nyukin_yotei_date': str(nyukin_date) if nyukin_date else '',
            'nyukin_yotei_text': nyukin_text,
            'is_collected':      is_collected,
        })

    dated   = sum(1 for p in payments if p['nyukin_yotei_date'])
    coll    = sum(1 for p in payments if p['is_collected'])
    print(f"  入金予定シート: {len(payments)}件（日付確定{dated}件, 受領済{coll}件）")
    return payments

# ===== メイン =====
def main():
    print("=" * 60)
    print("ワイズアシスト現場管理表 抽出パイプライン v2")
    print("(現場シートB4を売上計上一次源)")
    print("=" * 60)

    print("\n[1/5] Excel読込中...")
    wb_color = load_workbook(EXCEL_PATH, data_only=False)
    wb_data = load_workbook(EXCEL_PATH, data_only=True)

    print("[2/5] 進行管理シート抽出中...")
    sites_summary, excluded = extract_shinkou_kanri(wb_color, wb_data)
    print(f"  進行管理シート登録現場: {len(sites_summary)}件")

    # 進行管理に登録された現場のNo辞書
    sites_by_no = {s['genba_no']: s for s in sites_summary}

    # 現場シートにあるが進行管理にない現場も対象に加える
    site_sheet_names = sorted([s for s in wb_data.sheetnames if re.match(r'^F?\d+$', s)])
    extra_sheets = [s for s in site_sheet_names if s not in sites_by_no]
    print(f"  現場シートのみ(進行管理未登録): {len(extra_sheets)}件")

    print("[3/5] 各現場シート抽出 + B4パース中...")
    all_sites = []
    alerts = []  # (genba_no, kind, severity, message, detail)

    # シート名を正規化した逆引きマップ（全角→半角対応）
    sheet_norm_map = {_norm(name): name for name in wb_data.sheetnames}

    # まず進行管理ベースの現場
    for s in sites_summary:
        sheet_name = s['genba_no']
        actual_sheet = sheet_norm_map.get(_norm(sheet_name))  # 正規化マッチ
        if actual_sheet is not None:
            if actual_sheet != sheet_name:
                print(f"  [シート名正規化] {repr(sheet_name)} → {repr(actual_sheet)}")
            ws = wb_data[actual_sheet]
            template = identify_template(ws)
            detail = extract_site_sheet(ws, template)
            s.update(detail)
            s['template'] = template
            # B4をパース
            b4_raw = ws.cell(row=5, column=1).value
            a2_val = ws.cell(row=2, column=1).value
            b5_val = ws.cell(row=5, column=2).value
            chakkou = parse_date(a2_val)
            nyukin = parse_date(b5_val)
            chakkou_text = str(a2_val) if a2_val and not chakkou else ''
            nyukin_text = str(b5_val) if b5_val and not nyukin else ''
            p = parse_uriage_tsuki(b4_raw, chakkou, nyukin, chakkou_text, nyukin_text)
            s['b4_status'] = p['status']
            s['b4_raw'] = p.get('raw', '')
            s['b4_reason'] = p.get('reason', '')

            if p['status'] == 'ok':
                # B4ベースの按分を採用
                b4_allocations = [
                    {'year': y, 'month': m, 'ratio': r,
                     'sales_amount_zeikomi': s['juchu_zeikomi'] * r / 10,
                     'shitauke_amount_zeikomi': (s['shitauke_zeibetsu'] * 1.1 * r / 10
                                                  if s['sales_type'] == 'motouke' and s['shitauke_zeibetsu']
                                                  else None)}
                    for y, m, r in p['allocations']
                ]
                # 進行管理との差分を検出
                shinkou_set = {(a['year'], a['month'], round(a['ratio'], 2)) for a in s['allocations_shinkou']}
                b4_set = {(a['year'], a['month'], round(a['ratio'], 2)) for a in b4_allocations}
                if shinkou_set != b4_set:
                    # 月の集合自体が違うかどうか
                    shinkou_months = {(a['year'], a['month']) for a in s['allocations_shinkou']}
                    b4_months = {(a['year'], a['month']) for a in b4_allocations}
                    if shinkou_months != b4_months:
                        kind = '進行管理_月集合不一致'
                    else:
                        kind = '進行管理_割合不一致'
                    alerts.append({
                        'genba_no': sheet_name,
                        'kind': kind,
                        'severity': 'warning',
                        'message': '現場シートB4と進行管理シートで売上計上が異なる',
                        'detail': f'B4: {p["raw"]} / 進行管理: ' + ', '.join(
                            f'{a["year"]}/{a["month"]}={a["ratio"]:.1f}割' for a in s['allocations_shinkou'])
                    })
                # B4を採用
                s['allocations'] = b4_allocations
                if p.get('warnings'):
                    for w in p['warnings']:
                        alerts.append({'genba_no': sheet_name, 'kind': 'B4_警告',
                                       'severity': 'warning', 'message': w, 'detail': p['raw']})
            elif p['status'] == 'unparseable':
                # 進行管理ベースのまま
                alerts.append({'genba_no': sheet_name, 'kind': 'B4_パース不能',
                               'severity': 'warning',
                               'message': f'B4「{p.get("raw","")}」が解析できないため進行管理を使用',
                               'detail': p.get('reason', '')})
            else:  # empty
                if s['allocations_shinkou']:
                    alerts.append({'genba_no': sheet_name, 'kind': 'B4_空欄',
                                   'severity': 'info',
                                   'message': 'B4「売上計上月」が空欄、進行管理シートを採用',
                                   'detail': ''})
                else:
                    alerts.append({'genba_no': sheet_name, 'kind': 'B4_空欄',
                                   'severity': 'error',
                                   'message': 'B4と進行管理シート両方とも売上計上月が空欄',
                                   'detail': ''})
            all_sites.append(s)
        else:
            print(f"  [シートなし] genba_no={repr(sheet_name)} に対応するシートが見つかりません")
            s['template'] = 'NO_SHEET'
            s['costs'] = []
            s['b4_status'] = 'no_sheet'
            s['b4_raw'] = ''
            all_sites.append(s)

    # 現場シートのみ(進行管理未登録)の現場
    for sheet_name in extra_sheets:
        ws = wb_data[sheet_name]
        template = identify_template(ws)
        detail = extract_site_sheet(ws, template)
        # 進行管理由来の情報がないので、現場シートから受注額を取得
        juchu_zb = detail.get('juchu_zeibetsu') or 0
        # 受注税込 = 受注税別 × 1.1(または現場シートD8の値)
        juchu_zeikomi_from_sheet = detail.get('juchu_plus_tsuika') or (juchu_zb * 1.1)
        # 売上種別: 現場シートのR15「下請外注合計」があれば元請
        cb = detail.get('cost_breakdown', {})
        if template == 'motouke':
            sales_type = 'motouke'
            shitauke = cb.get('gaicyuhi', 0)
        else:
            sales_type = 'tsujou_kaitai'
            shitauke = None

        s = {
            'genba_no': sheet_name,
            'genba_name': detail.get('genba_name_sheet', ''),  # 進行管理由来がないのでシート名
            'juchu_zeikomi': juchu_zeikomi_from_sheet,
            'juchu_zeinuki': juchu_zb,
            'genka_mitsumori_shinkou': None,
            'hendouhi_shinkou': None,
            'shitauke_zeibetsu': shitauke,
            'sales_type': sales_type,
            'is_completed_shinkou': None,
            'chakkin_mikomi_text': '',
            'shiharai_text_shinkou': '',
            'sesshu_name': '',
            'shokaisha': '',
            'allocations_shinkou': [],
            'allocations': [],
            'template': template,
            **detail,
        }
        # B4パース
        b4_raw = ws.cell(row=5, column=1).value
        a2_val = ws.cell(row=2, column=1).value
        b5_val = ws.cell(row=5, column=2).value
        chakkou = parse_date(a2_val)
        nyukin = parse_date(b5_val)
        chakkou_text = str(a2_val) if a2_val and not chakkou else ''
        nyukin_text = str(b5_val) if b5_val and not nyukin else ''
        p = parse_uriage_tsuki(b4_raw, chakkou, nyukin, chakkou_text, nyukin_text)
        s['b4_status'] = p['status']
        s['b4_raw'] = p.get('raw', '')
        s['b4_reason'] = p.get('reason', '')

        if p['status'] == 'ok':
            b4_allocations = [
                {'year': y, 'month': m, 'ratio': r,
                 'sales_amount_zeikomi': juchu_zeikomi_from_sheet * r / 10,
                 'shitauke_amount_zeikomi': (shitauke * 1.1 * r / 10
                                              if sales_type == 'motouke' and shitauke
                                              else None)}
                for y, m, r in p['allocations']
            ]
            s['allocations'] = b4_allocations
            alerts.append({'genba_no': sheet_name, 'kind': '進行管理_未登録',
                           'severity': 'info',
                           'message': '進行管理シートに未登録(現場シートのみ)',
                           'detail': f'B4: {p["raw"]}'})
            if p.get('warnings'):
                for w in p['warnings']:
                    alerts.append({'genba_no': sheet_name, 'kind': 'B4_警告',
                                   'severity': 'warning', 'message': w, 'detail': p['raw']})
        elif p['status'] == 'unparseable':
            alerts.append({'genba_no': sheet_name, 'kind': 'B4_パース不能',
                           'severity': 'error',
                           'message': f'B4「{p.get("raw","")}」が解析できず、進行管理にも未登録',
                           'detail': p.get('reason', '')})
        else:  # empty
            alerts.append({'genba_no': sheet_name, 'kind': '進行管理_未登録_B4空',
                           'severity': 'error',
                           'message': '進行管理に未登録、かつB4も空欄。売上計上月不明',
                           'detail': ''})
        all_sites.append(s)

    print(f"  抽出現場合計: {len(all_sites)}件")
    print(f"  検出アラート: {len(alerts)}件")

    print("[3.5/5] 入金予定シート抽出中...")
    payments = extract_nyukin_yotei(wb_data)

    print("[4/5] SQLite保存中...")
    save_to_sqlite(all_sites, excluded, alerts, payments)

    print("[5/5] 完了")
    return all_sites, excluded, alerts

def save_to_sqlite(sites, excluded, alerts=None, payments=None):
    # cf_actual_balance（手入力残高）を退避してDB再作成後に復元
    cf_actual_backup = []
    if Path(DB_PATH).exists():
        try:
            _tmp = sqlite3.connect(DB_PATH)
            cf_actual_backup = _tmp.execute(
                "SELECT ym, actual_balance, memo FROM cf_actual_balance").fetchall()
            _tmp.close()
        except Exception:
            pass

    Path(DB_PATH).unlink(missing_ok=True)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute("""CREATE TABLE sites (
        genba_no TEXT PRIMARY KEY,
        genba_name TEXT, hatchusha TEXT, ichijuke TEXT,
        sesshu_name TEXT, shokaisha TEXT,
        chakkou_date TEXT, sales_type TEXT, koji_kubun TEXT,
        template TEXT, sekoujoukyou TEXT, is_completed INTEGER,
        juchu_zeikomi REAL, juchu_zeinuki REAL, juchu_zeibetsu REAL, shohizei REAL,
        tsuika_koji_zeibetsu REAL,
        genka_mitsumori_shinkou REAL, hendouhi_shinkou REAL,
        shitauke_zeibetsu REAL,
        genka_yotei REAL, genka_jissai REAL,
        rieki_ritsu_yotei REAL, rieki_ritsu_jissai REAL,
        rieki_yotei REAL, rieki_jissai REAL,
        nyukin_yotei_date TEXT,
        chakkin_mikomi_text TEXT, shiharai_text_shinkou TEXT,
        uriage_tsuki_text TEXT,
        b4_status TEXT, b4_raw TEXT
    )""")
    c.execute("""CREATE TABLE sales_allocations (
        genba_no TEXT, year INTEGER, month INTEGER,
        ratio_wari REAL, sales_amount_zeikomi REAL,
        shitauke_amount_zeikomi REAL, source TEXT,
        PRIMARY KEY(genba_no, year, month)
    )""")
    c.execute("""CREATE TABLE sales_allocations_shinkou (
        genba_no TEXT, year INTEGER, month INTEGER,
        ratio_wari REAL, sales_amount_zeikomi REAL,
        shitauke_amount_zeikomi REAL,
        PRIMARY KEY(genba_no, year, month)
    )""")
    c.execute("""CREATE TABLE costs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        genba_no TEXT, hi_moku TEXT, torihiki_saki TEXT,
        hassei_date TEXT, shiharai_date TEXT, hassei_tsuki TEXT,
        kingaku_zeibetsu REAL, kingaku_zeikomi REAL, memo TEXT
    )""")
    c.execute("""CREATE TABLE excluded_sites (
        genba_no TEXT, genba_name TEXT, reason TEXT
    )""")
    c.execute("""CREATE TABLE alerts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        genba_no TEXT, kind TEXT, severity TEXT, message TEXT, detail TEXT
    )""")
    c.execute("""CREATE TABLE site_payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        genba_no TEXT, genba_name TEXT, seikyu_kubun TEXT,
        kingaku_zeikomi REAL,
        nyukin_yotei_date TEXT,
        nyukin_yotei_text TEXT,
        is_collected INTEGER DEFAULT 0
    )""")
    c.execute("""CREATE TABLE cf_actual_balance (
        ym TEXT PRIMARY KEY, actual_balance REAL, memo TEXT DEFAULT ''
    )""")
    for row in cf_actual_backup:
        c.execute("INSERT OR REPLACE INTO cf_actual_balance VALUES (?,?,?)", row)

    for s in sites:
        # 完了判定: 進行管理優先、なければ現場シートの施工状況
        is_completed = s.get('is_completed_shinkou')
        if is_completed is None:
            sk = str(s.get('sekoujoukyou') or '')
            if '完了' in sk:
                is_completed = True
            else:
                is_completed = False

        c.execute("""INSERT INTO sites VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
            s['genba_no'], s.get('genba_name'),
            str(s.get('hatchusha') or ''), str(s.get('ichijuke') or ''),
            str(s.get('sesshu_name') or ''), str(s.get('shokaisha') or ''),
            str(s.get('chakkou_date') or ''), s['sales_type'],
            str(s.get('koji_kubun') or ''), s.get('template'),
            str(s.get('sekoujoukyou') or ''),
            1 if is_completed else 0,
            s.get('juchu_zeikomi'), s.get('juchu_zeinuki'), s.get('juchu_zeibetsu'), s.get('shohizei'),
            s.get('tsuika_koji_zeibetsu'),
            s.get('genka_mitsumori_shinkou'), s.get('hendouhi_shinkou'),
            s.get('shitauke_zeibetsu'),
            s.get('genka_yotei'), s.get('genka_jissai'),
            s.get('rieki_ritsu_yotei'), s.get('rieki_ritsu_jissai'),
            s.get('rieki_yotei'), s.get('rieki_jissai'),
            str(s.get('nyukin_yotei_date') or ''),
            str(s.get('chakkin_mikomi_text') or ''),
            str(s.get('shiharai_text_shinkou') or ''),
            str(s.get('uriage_tsuki_text') or ''),
            s.get('b4_status', ''), s.get('b4_raw', ''),
        ))
        for a in s.get('allocations', []):
            c.execute("INSERT OR REPLACE INTO sales_allocations VALUES (?,?,?,?,?,?,?)", (
                s['genba_no'], a['year'], a['month'], a['ratio'],
                a['sales_amount_zeikomi'], a.get('shitauke_amount_zeikomi'),
                'B4' if s.get('b4_status') == 'ok' else '進行管理'
            ))
        for a in s.get('allocations_shinkou', []):
            c.execute("INSERT OR REPLACE INTO sales_allocations_shinkou VALUES (?,?,?,?,?,?)", (
                s['genba_no'], a['year'], a['month'], a['ratio'],
                a['sales_amount_zeikomi'], a.get('shitauke_amount_zeikomi')
            ))
        for cost in s.get('costs', []):
            c.execute("INSERT INTO costs (genba_no, hi_moku, torihiki_saki, hassei_date, shiharai_date, hassei_tsuki, kingaku_zeibetsu, kingaku_zeikomi, memo) VALUES (?,?,?,?,?,?,?,?,?)", (
                s['genba_no'], cost['hi_moku'],
                cost.get('torihiki_saki', ''),
                str(cost.get('hassei_date') or ''),
                str(cost.get('shiharai_date') or ''),
                cost.get('hassei_tsuki', ''),
                cost.get('kingaku_zeibetsu'),
                cost.get('kingaku_zeikomi'),
                cost.get('memo', '')
            ))
    for ex in excluded:
        c.execute("INSERT INTO excluded_sites VALUES (?,?,?)", (ex[0], str(ex[1] or ''), ex[2]))

    for a in (alerts or []):
        c.execute("INSERT INTO alerts (genba_no, kind, severity, message, detail) VALUES (?,?,?,?,?)",
                  (a['genba_no'], a['kind'], a['severity'], a['message'], a.get('detail', '')))

    for p in (payments or []):
        c.execute("""INSERT INTO site_payments
                     (genba_no, genba_name, seikyu_kubun, kingaku_zeikomi,
                      nyukin_yotei_date, nyukin_yotei_text, is_collected)
                     VALUES (?,?,?,?,?,?,?)""",
                  (p['genba_no'], p['genba_name'], p['seikyu_kubun'],
                   p['kingaku_zeikomi'], p['nyukin_yotei_date'],
                   p['nyukin_yotei_text'], p['is_collected']))

    conn.commit()

    print(f"  sites: {c.execute('SELECT COUNT(*) FROM sites').fetchone()[0]}件")
    print(f"  sales_allocations: {c.execute('SELECT COUNT(*) FROM sales_allocations').fetchone()[0]}件")
    print(f"  costs: {c.execute('SELECT COUNT(*) FROM costs').fetchone()[0]}件")
    print(f"  site_payments: {c.execute('SELECT COUNT(*) FROM site_payments').fetchone()[0]}件")
    print(f"  alerts: {c.execute('SELECT COUNT(*) FROM alerts').fetchone()[0]}件")
    conn.close()

if __name__ == '__main__':
    main()
